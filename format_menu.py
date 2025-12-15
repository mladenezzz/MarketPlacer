# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Читаем исходный файл
df = pd.read_excel('z:/menu_diet.xlsx', header=0)
df.columns = ['День', 'Прием', 'Блюдо', 'Порция_г', 'Белок_г']

# Создаем новый Excel файл
wb = Workbook()
ws = wb.active
ws.title = 'Меню на неделю'

# Стили
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
day_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
meal1_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
meal2_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
header_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
white_font = Font(bold=True, size=11, color='FFFFFF')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Заголовок
ws.merge_cells('A1:E1')
ws['A1'] = 'МЕНЮ НА НЕДЕЛЮ (диетическое питание)'
ws['A1'].font = Font(bold=True, size=16)
ws['A1'].alignment = center

row = 3

# Заголовок таблицы
headers = ['День', 'Приём', 'Блюдо', 'Порция (г)', 'Белок (г)']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = Font(bold=True, size=11)
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center
row += 1

days = df['День'].unique()
prev_day = None

for day in days:
    day_data = df[df['День'] == day]
    day_rows = len(day_data)
    start_row = row

    for meal_num in [1, 2]:
        meal_data = day_data[day_data['Прием'] == meal_num]
        if len(meal_data) == 0:
            continue

        meal_start_row = row
        meal_fill = meal1_fill if meal_num == 1 else meal2_fill
        meal_name = 'Завтрак' if meal_num == 1 else 'Ужин'

        for idx, (_, item) in enumerate(meal_data.iterrows()):
            # Блюдо
            cell = ws.cell(row=row, column=3, value=item['Блюдо'])
            cell.border = thin_border
            cell.alignment = left
            cell.fill = meal_fill

            # Порция
            cell = ws.cell(row=row, column=4, value=item['Порция_г'])
            cell.border = thin_border
            cell.alignment = center
            cell.fill = meal_fill

            # Белок
            cell = ws.cell(row=row, column=5, value=item['Белок_г'])
            cell.border = thin_border
            cell.alignment = center
            cell.fill = meal_fill

            row += 1

        # Объединяем ячейки приёма пищи
        if row > meal_start_row:
            ws.merge_cells(f'B{meal_start_row}:B{row-1}')
            cell = ws.cell(row=meal_start_row, column=2, value=meal_name)
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border
            cell.alignment = center
            cell.fill = meal_fill
            # Добавляем границы для объединённых ячеек
            for r in range(meal_start_row, row):
                ws.cell(row=r, column=2).border = thin_border

    # Объединяем ячейки дня
    if row > start_row:
        ws.merge_cells(f'A{start_row}:A{row-1}')
        cell = ws.cell(row=start_row, column=1, value=day)
        cell.font = Font(bold=True, size=11)
        cell.border = thin_border
        cell.alignment = center
        cell.fill = day_fill
        cell.font = white_font
        # Добавляем границы для объединённых ячеек
        for r in range(start_row, row):
            ws.cell(row=r, column=1).border = thin_border

# Ширина колонок
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12

# Настройки печати
ws.print_title_rows = '3:3'  # Повторять заголовок на каждой странице
ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

wb.save('z:/menu_diet_print.xlsx')
print('Файл успешно сохранен: z:/menu_diet_print.xlsx')
