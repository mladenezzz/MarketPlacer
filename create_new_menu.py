# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Меню для мужчины 40 лет, 190 см, 105 кг (малая активность)
# Цель: ~1700-1900 ккал, 120-140г белка, дефицит калорий
# 2 приёма пищи: 12:00 и 18:00

menu_data = [
    # Понедельник
    ("Понедельник", 1, "Яйца (3 целых + 2 белка)", 230, 30, 280, 16, 2),
    ("Понедельник", 1, "Куриная грудка запечённая", 200, 44, 220, 3, 0),
    ("Понедельник", 1, "Гречка", 100, 12, 350, 4, 68),
    ("Понедельник", 1, "Салат (огурцы+помидоры+зелень+масло)", 250, 2, 100, 7, 8),

    ("Понедельник", 2, "Рыба (треска) запечённая с лимоном", 300, 54, 245, 2, 1),
    ("Понедельник", 2, "Квашеная капуста", 150, 2, 30, 0, 5),
    ("Понедельник", 2, "Творог 5%", 300, 42, 310, 6, 9),
    ("Понедельник", 2, "Орехи (миндаль)", 30, 6, 180, 16, 6),

    # Вторник
    ("Вторник", 1, "Омлет (4 яйца + молоко) с зеленью", 300, 28, 345, 24, 4),
    ("Вторник", 1, "Печень куриная тушёная", 200, 40, 280, 12, 2),
    ("Вторник", 1, "Рис бурый", 100, 8, 350, 3, 72),
    ("Вторник", 1, "Овощи тушёные", 250, 4, 90, 3, 15),

    ("Вторник", 2, "Куриная грудка отварная", 250, 55, 275, 3, 0),
    ("Вторник", 2, "Салат (шпинат+огурец+зелень+масло)", 300, 4, 110, 7, 8),
    ("Вторник", 2, "Творог 5%", 200, 28, 210, 4, 6),
    ("Вторник", 2, "Орехи грецкие", 25, 4, 165, 16, 3),
    ("Вторник", 2, "Кефир 1%", 300, 9, 120, 3, 12),

    # Среда
    ("Среда", 1, "Творог 5%", 250, 35, 260, 5, 8),
    ("Среда", 1, "Яйца варёные (3 шт)", 180, 20, 230, 15, 2),
    ("Среда", 1, "Индейка запечённая", 220, 48, 250, 3, 0),
    ("Среда", 1, "Гречка", 100, 12, 350, 4, 68),
    ("Среда", 1, "Салат (листовой+огурцы+зелень+масло)", 250, 2, 90, 7, 6),

    ("Среда", 2, "Лосось запечённый с лимоном", 220, 44, 445, 28, 1),
    ("Среда", 2, "Квашеная капуста", 150, 2, 30, 0, 5),
    ("Среда", 2, "Авокадо", 70, 1, 110, 10, 6),
    ("Среда", 2, "Кефир 1%", 400, 12, 160, 4, 16),

    # Четверг
    ("Четверг", 1, "Яичница (4 яйца) с зеленью", 240, 26, 325, 24, 2),
    ("Четверг", 1, "Куриные бёдра без кожи", 220, 40, 310, 11, 0),
    ("Четверг", 1, "Гречка", 100, 12, 350, 4, 68),
    ("Четверг", 1, "Салат (огурец+помидор+зелень)", 250, 2, 50, 0, 9),

    ("Четверг", 2, "Печень говяжья тушёная", 200, 40, 250, 8, 5),
    ("Четверг", 2, "Квашеная капуста", 150, 2, 30, 0, 5),
    ("Четверг", 2, "Творог 5%", 250, 35, 260, 5, 8),
    ("Четверг", 2, "Орехи (миндаль)", 25, 5, 150, 13, 5),
    ("Четверг", 2, "Кефир 1%", 300, 9, 120, 3, 12),

    # Пятница
    ("Пятница", 1, "Омлет (4 яйца + молоко) с зеленью", 300, 28, 345, 24, 4),
    ("Пятница", 1, "Рыба (минтай) на пару с лимоном", 300, 48, 215, 2, 1),
    ("Пятница", 1, "Рис бурый", 100, 8, 350, 3, 72),
    ("Пятница", 1, "Салат (шпинат+огурец+зелень)", 200, 3, 40, 0, 6),

    ("Пятница", 2, "Индейка запечённая", 250, 55, 290, 4, 0),
    ("Пятница", 2, "Квашеная капуста", 150, 2, 30, 0, 5),
    ("Пятница", 2, "Авокадо", 70, 1, 110, 10, 6),
    ("Пятница", 2, "Творог 5%", 250, 35, 260, 5, 8),
    ("Пятница", 2, "Кефир 1%", 300, 9, 120, 3, 12),

    # Суббота
    ("Суббота", 1, "Яйца (3 целых + 2 белка)", 230, 30, 280, 16, 2),
    ("Суббота", 1, "Свинина нежирная запечённая", 220, 44, 375, 20, 0),
    ("Суббота", 1, "Рис бурый", 100, 8, 350, 3, 72),
    ("Суббота", 1, "Салат (листовой+огурцы+зелень+масло)", 250, 2, 90, 7, 6),

    ("Суббота", 2, "Куриная грудка запечённая", 250, 55, 275, 3, 0),
    ("Суббота", 2, "Овощи гриль (кабачок, перец)", 300, 4, 100, 4, 15),
    ("Суббота", 2, "Творог 5%", 250, 35, 260, 5, 8),
    ("Суббота", 2, "Орехи грецкие", 25, 4, 165, 16, 3),
    ("Суббота", 2, "Кефир 1%", 300, 9, 120, 3, 12),

    # Воскресенье
    ("Воскресенье", 1, "Омлет (4 яйца) с зеленью", 280, 26, 315, 22, 3),
    ("Воскресенье", 1, "Говядина стейк", 250, 60, 450, 20, 0),
    ("Воскресенье", 1, "Гречка", 100, 12, 350, 4, 68),
    ("Воскресенье", 1, "Салат (шпинат+огурец+помидор+масло)", 250, 3, 80, 5, 8),

    ("Воскресенье", 2, "Рыба (треска) запечённая с лимоном", 300, 54, 245, 2, 1),
    ("Воскресенье", 2, "Квашеная капуста", 150, 2, 30, 0, 5),
    ("Воскресенье", 2, "Авокадо", 70, 1, 110, 10, 6),
    ("Воскресенье", 2, "Творог 5%", 250, 35, 260, 5, 8),
    ("Воскресенье", 2, "Кефир 1%", 300, 9, 120, 3, 12),
]

# Создаём DataFrame
df = pd.DataFrame(menu_data, columns=['День', 'Прием', 'Блюдо', 'Порция_г', 'Белок_г', 'Ккал', 'Жиры_г', 'Углеводы_г'])

# Создаем Excel файл
wb = Workbook()
ws = wb.active
ws.title = 'Меню на неделю'

# Стили
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
day_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
meal1_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # Завтрак - зелёный
meal2_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # Обед - оранжевый
meal3_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')  # Ужин - голубой
total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Итого - жёлтый
header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
white_font = Font(bold=True, size=11, color='FFFFFF')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
right = Alignment(horizontal='right', vertical='center')

# Заголовок
ws.merge_cells('A1:H1')
ws['A1'] = 'МЕНЮ НА НЕДЕЛЮ — Мужчина 40 лет, 190 см, 105 кг (дефицит калорий)'
ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = center

ws.merge_cells('A2:H2')
ws['A2'] = 'Цель: ~1700-1900 ккал, 120-140г белка | 2 приёма: 12:00 и 18:00'
ws['A2'].font = Font(italic=True, size=11)
ws['A2'].alignment = center

row = 4

# Заголовок таблицы
headers = ['День', 'Приём', 'Блюдо', 'Порция (г)', 'Белок (г)', 'Ккал', 'Жиры (г)', 'Углев. (г)']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.font = Font(bold=True, size=10)
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center
row += 1

days = df['День'].unique()
meal_names = {1: '12:00', 2: '18:00'}
meal_fills = {1: meal1_fill, 2: meal2_fill}

for day in days:
    day_data = df[df['День'] == day]
    start_row = row

    day_totals = {'Белок_г': 0, 'Ккал': 0, 'Жиры_г': 0, 'Углеводы_г': 0}

    for meal_num in [1, 2]:
        meal_data = day_data[day_data['Прием'] == meal_num]
        if len(meal_data) == 0:
            continue

        meal_start_row = row
        meal_fill = meal_fills[meal_num]
        meal_name = meal_names[meal_num]

        for _, item in meal_data.iterrows():
            # Блюдо
            cell = ws.cell(row=row, column=3, value=item['Блюдо'])
            cell.border = thin_border
            cell.alignment = left
            cell.fill = meal_fill

            # Порция
            cell = ws.cell(row=row, column=4, value=item['Порция_г'])
            cell.border = thin_border
            cell.alignment = center
            cell.fill = meal_fill

            # Белок
            cell = ws.cell(row=row, column=5, value=item['Белок_г'])
            cell.border = thin_border
            cell.alignment = center
            cell.fill = meal_fill
            day_totals['Белок_г'] += item['Белок_г']

            # Ккал
            cell = ws.cell(row=row, column=6, value=item['Ккал'])
            cell.border = thin_border
            cell.alignment = center
            cell.fill = meal_fill
            day_totals['Ккал'] += item['Ккал']

            # Жиры
            cell = ws.cell(row=row, column=7, value=item['Жиры_г'])
            cell.border = thin_border
            cell.alignment = center
            cell.fill = meal_fill
            day_totals['Жиры_г'] += item['Жиры_г']

            # Углеводы
            cell = ws.cell(row=row, column=8, value=item['Углеводы_г'])
            cell.border = thin_border
            cell.alignment = center
            cell.fill = meal_fill
            day_totals['Углеводы_г'] += item['Углеводы_г']

            row += 1

        # Объединяем ячейки приёма пищи
        if row > meal_start_row:
            ws.merge_cells(f'B{meal_start_row}:B{row-1}')
            cell = ws.cell(row=meal_start_row, column=2, value=meal_name)
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border
            cell.alignment = center
            cell.fill = meal_fill
            for r in range(meal_start_row, row):
                ws.cell(row=r, column=2).border = thin_border

    # Строка итогов за день
    ws.merge_cells(f'C{row}:D{row}')
    cell = ws.cell(row=row, column=3, value='ИТОГО за день:')
    cell.font = Font(bold=True, size=10)
    cell.fill = total_fill
    cell.border = thin_border
    cell.alignment = right
    ws.cell(row=row, column=4).border = thin_border
    ws.cell(row=row, column=4).fill = total_fill

    for col, key in [(5, 'Белок_г'), (6, 'Ккал'), (7, 'Жиры_г'), (8, 'Углеводы_г')]:
        cell = ws.cell(row=row, column=col, value=day_totals[key])
        cell.font = Font(bold=True, size=10)
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = center

    row += 1

    # Объединяем ячейки дня (включая строку итогов)
    ws.merge_cells(f'A{start_row}:A{row-1}')
    cell = ws.cell(row=start_row, column=1, value=day)
    cell.font = white_font
    cell.border = thin_border
    cell.alignment = center
    cell.fill = day_fill
    for r in range(start_row, row):
        ws.cell(row=r, column=1).border = thin_border

# Ширина колонок
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 32
ws.column_dimensions['D'].width = 11
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 10

# Настройки печати
ws.print_title_rows = '4:4'
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

wb.save('z:/menu_diet_new.xlsx')
print('Новое меню сохранено: z:/menu_diet_new.xlsx')

# Выводим средние показатели
print("\n=== СРЕДНИЕ ПОКАЗАТЕЛИ ЗА ДЕНЬ ===")
daily_totals = df.groupby('День').agg({
    'Белок_г': 'sum',
    'Ккал': 'sum',
    'Жиры_г': 'sum',
    'Углеводы_г': 'sum'
}).mean()
print(f"Калории: {daily_totals['Ккал']:.0f} ккал")
print(f"Белок: {daily_totals['Белок_г']:.0f} г")
print(f"Жиры: {daily_totals['Жиры_г']:.0f} г")
print(f"Углеводы: {daily_totals['Углеводы_г']:.0f} г")
